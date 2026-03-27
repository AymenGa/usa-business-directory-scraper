"""
build_excel.py
Builds output/results.xlsx with a professional template.
- Cover sheet with project summary and clickable navigation
- TWO sheets per completed category:
    {Category} - Directory    (Yelp businesses + Etsy shop profiles, 10 columns)
    {Category} - Marketplace  (Etsy products only, 8 columns)
- Reads extraction dates from output/categories/metadata.json

Source files per category (in output/categories/):
  {slug}_directory.xlsx    → Etsy shops (from etsy_shops.py) + Yelp businesses (TBD)
  {slug}_marketplace.xlsx  → Etsy products (from etsy_products.py)

Run: python scraper/build_excel.py
"""

import os
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CATEGORIES_DIR = os.path.join(BASE_DIR, "output", "categories")
OUTPUT_FILE    = os.path.join(BASE_DIR, "output", "results.xlsx")
METADATA_FILE  = os.path.join(CATEGORIES_DIR, "metadata.json")

# ── Category config ───────────────────────────────────────────────────────────
CATEGORIES = [
    {"name": "Fashion / Accessories",  "color": "C0392B"},
    {"name": "Electronics / Gadgets",  "color": "2980B9"},
    {"name": "Home and Garden",         "color": "27AE60"},
    {"name": "Toys and Games",          "color": "F39C12"},
    {"name": "Health and Beauty",       "color": "8E44AD"},
    {"name": "Digital Products",        "color": "16A085"},
    {"name": "Automotive",              "color": "2C3E50"},
    {"name": "Food and Groceries",      "color": "E67E22"},
    {"name": "Sports and Outdoors",     "color": "1ABC9C"},
    {"name": "Books and Media",         "color": "D35400"},
    {"name": "Arts and Crafts",         "color": "E91E63"},
    {"name": "Professional Services",   "color": "546E7A"},
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def white_bold_font(size=11):
    return Font(bold=True, color="FFFFFF", size=size)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def autofit_columns(ws, min_width=12, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))

def slug(name):
    return name.lower().replace(" / ", "_").replace(" & ", "_").replace(" ", "_")

def clean_price(value):
    """Remove $ and commas -> plain number."""
    if value is None:
        return value
    s = str(value).replace("$", "").replace(",", "").strip()
    if "." in s:
        return float(s)
    return int(s)

# ── Metadata ──────────────────────────────────────────────────────────────────
def load_metadata():
    if os.path.exists(METADATA_FILE):
        with open(METADATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_metadata(meta):
    os.makedirs(os.path.dirname(METADATA_FILE), exist_ok=True)
    with open(METADATA_FILE, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

# ── Load data files ──────────────────────────────────────────────────────────
def load_xlsx(path):
    """Load an xlsx file. Returns (headers, data) or (None, None)."""
    if not os.path.exists(path):
        return None, None

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None, None

    headers = list(rows[0])
    data = list(rows[1:])
    return headers, data


def load_directory_data(category_name):
    """Load {slug}_directory.xlsx. Returns (headers, data)."""
    s = slug(category_name)
    path = os.path.join(CATEGORIES_DIR, f"{s}_directory.xlsx")
    return load_xlsx(path)


def load_marketplace_data(category_name):
    """Load {slug}_marketplace.xlsx. Returns (headers, data)."""
    s = slug(category_name)
    path = os.path.join(CATEGORIES_DIR, f"{s}_marketplace.xlsx")
    return load_xlsx(path)


# ── Write a styled sheet ─────────────────────────────────────────────────────
def write_sheet(wb, sheet_title, tab_color, headers, data):
    """Write a styled sheet. Returns the worksheet."""
    safe_title = sheet_title[:31]
    ws = wb.create_sheet(title=safe_title)
    ws.sheet_properties.tabColor = tab_color

    header_fill = hex_fill(tab_color)
    alt_fill    = hex_fill("F2F2F2")

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = white_bold_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 20

    # Detect special columns by name
    price_col  = None
    rating_col = None
    for i, h in enumerate(headers, start=1):
        hl = str(h).lower()
        if hl == "price":
            price_col = i
        if "star" in hl or "rating" in hl:
            rating_col = i

    # Data rows
    for row_idx, row in enumerate(data, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            if col_idx == price_col and value not in (None, "N/A", ""):
                try:
                    value = clean_price(value)
                except Exception:
                    pass
            if col_idx == rating_col:
                if value in (None, "", "N/A"):
                    value = "N/A"
                else:
                    try:
                        value = float(value)
                    except Exception:
                        value = "N/A"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = thin_border()

            if col_idx == price_col and isinstance(value, (int, float)):
                cell.number_format = '$#,##0.00' if isinstance(value, float) else '$#,##0'
            if col_idx == rating_col and isinstance(value, float):
                cell.number_format = '0.00 "★"'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)
    return ws


# ── Cover sheet ───────────────────────────────────────────────────────────────
def write_cover_sheet(wb, summary):
    """Write the cover/overview sheet.
    summary is a list of tuples:
      (category, status, dir_count, mkt_count, dir_sheet, mkt_sheet, date_extracted, source)
    """
    ws = wb.create_sheet(title="Overview", index=0)
    ws.sheet_properties.tabColor = "2C3E50"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value     = "USA Business Directory & Marketplace"
    title_cell.font      = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill      = hex_fill("2C3E50")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value     = f"Client: Joseph Clay  |  Generated: {date.today().strftime('%B %d, %Y')}"
    sub.font      = Font(italic=True, size=10, color="FFFFFF")
    sub.fill      = hex_fill("34495E")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Spacer
    ws.row_dimensions[3].height = 10

    # Table headers
    col_headers = ["Category", "Status", "Directory Rows", "Marketplace Rows",
                   "Date Extracted", "Source", "", "Go to Directory", "Go to Marketplace"]
    header_fill = hex_fill("2C3E50")
    for col_idx, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = white_bold_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
    # Merge "Source" header across two columns (F4:G4)
    ws.merge_cells("F4:G4")
    ws.row_dimensions[4].height = 20

    alt_fill = hex_fill("F2F2F2")

    for row_idx, entry in enumerate(summary, start=5):
        cat, status, dir_count, mkt_count, dir_sheet, mkt_sheet, date_extracted, source = entry
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()

        # Category name
        c = ws.cell(row=row_idx, column=1, value=cat)
        c.fill = fill; c.font = Font(bold=True, size=10); c.alignment = Alignment(vertical="center"); c.border = thin_border()

        # Status
        c = ws.cell(row=row_idx, column=2, value=status)
        c.fill = fill; c.font = Font(color="27AE60" if status == "Done" else "E67E22", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()

        # Directory count
        c = ws.cell(row=row_idx, column=3, value=dir_count if dir_count else "-")
        c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()

        # Marketplace count
        c = ws.cell(row=row_idx, column=4, value=mkt_count if mkt_count else "-")
        c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()

        # Date Extracted
        c = ws.cell(row=row_idx, column=5, value=date_extracted if date_extracted else "-")
        c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
        c.font = Font(size=10, color="555555" if date_extracted else "AAAAAA")

        # Source websites — split into separate clickable columns
        sources = [s.strip() for s in source.split(",")] if source else []
        dir_source = sources[1] if len(sources) > 1 else (sources[0] if sources else "")
        mkt_source = sources[0] if sources else ""

        # Directory Source (column 6)
        c = ws.cell(row=row_idx, column=6, value=dir_source if dir_source else "-")
        c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
        if dir_source:
            url = dir_source if dir_source.startswith("http") else f"https://{dir_source}"
            c.hyperlink = url
            c.font = Font(size=10, color="2980B9", underline="single")
        else:
            c.font = Font(size=10, color="AAAAAA")

        # Marketplace Source (column 7)
        c = ws.cell(row=row_idx, column=7, value=mkt_source if mkt_source else "-")
        c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
        if mkt_source:
            url = mkt_source if mkt_source.startswith("http") else f"https://{mkt_source}"
            c.hyperlink = url
            c.font = Font(size=10, color="2980B9", underline="single")
        else:
            c.font = Font(size=10, color="AAAAAA")

        # Navigation link — Directory (column 8)
        all_sheets = [s.title for s in wb.worksheets]
        c = ws.cell(row=row_idx, column=8)
        c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
        if dir_sheet and dir_sheet in all_sheets:
            c.value = "Open Directory"
            c.hyperlink = f"#'{dir_sheet}'!A1"
            c.font = Font(color="2980B9", underline="single", size=10)
        else:
            c.value = "-"
            c.font = Font(color="AAAAAA", italic=True, size=10)

        # Navigation link — Marketplace (column 9)
        c = ws.cell(row=row_idx, column=9)
        c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border()
        if mkt_sheet and mkt_sheet in all_sheets:
            c.value = "Open Marketplace"
            c.hyperlink = f"#'{mkt_sheet}'!A1"
            c.font = Font(color="2980B9", underline="single", size=10)
        else:
            c.value = "-"
            c.font = Font(color="AAAAAA", italic=True, size=10)

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A5"


# ── Main ──────────────────────────────────────────────────────────────────────
def build():
    meta = load_metadata()
    save_metadata(meta)

    wb = Workbook()
    del wb[wb.sheetnames[0]]

    summary = []

    for cat in CATEGORIES:
        name  = cat["name"]
        color = cat["color"]
        source = meta.get(name, {}).get("source")
        date_extracted = meta.get(name, {}).get("date_extracted")

        dir_headers, dir_data = load_directory_data(name)
        mkt_headers, mkt_data = load_marketplace_data(name)

        dir_sheet_name = None
        mkt_sheet_name = None
        dir_count = 0
        mkt_count = 0

        # Write Directory sheet
        if dir_headers and dir_data:
            title = f"{name.replace('/', '-')} - Directory"[:31]
            ws = write_sheet(wb, title, color, dir_headers, dir_data)
            dir_sheet_name = ws.title
            dir_count = len(dir_data)

        # Write Marketplace sheet
        if mkt_headers and mkt_data:
            title = f"{name.replace('/', '-')} - Marketplace"[:31]
            ws = write_sheet(wb, title, color, mkt_headers, mkt_data)
            mkt_sheet_name = ws.title
            mkt_count = len(mkt_data)

        status = "Done" if (dir_data or mkt_data) else "Pending"
        summary.append((name, status, dir_count, mkt_count, dir_sheet_name,
                         mkt_sheet_name, date_extracted, source))

    write_cover_sheet(wb, summary)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"[OK] Saved: {OUTPUT_FILE}")
    for cat, status, d_count, m_count, *_ in summary:
        mark = "[DONE]   " if status == "Done" else "[PENDING]"
        d = f"{d_count:>5} dir" if d_count else "    -    "
        m = f"{m_count:>5} mkt" if m_count else "    -    "
        print(f"  {mark}  {cat:<28} {d}  |  {m}")


if __name__ == "__main__":
    build()
