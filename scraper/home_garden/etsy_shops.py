"""
Home and Garden — Etsy shop profile enricher (directory data).
Visits each unique shop page to extract location, rating, icon.
Creates Directory rows (10 columns) from Etsy shop profiles.

Run:
  python scraper/home_garden/etsy_shops.py --test   # first 10 shops
  python scraper/home_garden/etsy_shops.py          # all shops
"""

import argparse
import json
import os
import time
import random
import logging
import openpyxl
from bs4 import BeautifulSoup

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CATEGORIES_DIR = os.path.join(BASE_DIR, "output", "categories")
METADATA_FILE  = os.path.join(CATEGORIES_DIR, "metadata.json")

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Home and Garden config ───────────────────────────────────────────────────
CATEGORY_NAME = "Home and Garden"
CATEGORY_SLUG = "home_and_garden"
SOURCE        = "www.etsy.com"


# ── Checkpoint path ───────────────────────────────────────────────────────────
CHECKPOINT_FILE = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_shops_checkpoint.json")
CHECKPOINT_EVERY = 10


def save_checkpoint(directory_rows, completed_shops):
    os.makedirs(CATEGORIES_DIR, exist_ok=True)
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump({"rows": directory_rows, "completed": list(completed_shops)}, f, ensure_ascii=False)
    log.info(f"  Checkpoint saved ({len(directory_rows)} rows, {len(completed_shops)} shops done)")


def load_checkpoint():
    if not os.path.exists(CHECKPOINT_FILE):
        return [], set()
    with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
        payload = json.load(f)
    rows = payload.get("rows", [])
    completed = set(payload.get("completed", []))
    log.info(f"Resuming from checkpoint: {len(rows)} rows, {len(completed)} shops already done")
    return rows, completed


# ── Browser helpers ──────────────────────────────────────────────────────────
def create_driver():
    import undetected_chromedriver as uc
    options = uc.ChromeOptions()
    options.add_argument("--lang=en-US")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = uc.Chrome(options=options, headless=False, version_main=146)
    driver.set_page_load_timeout(30)
    return driver


def is_driver_alive(driver):
    try:
        _ = driver.title
        return True
    except Exception:
        return False


def fetch_page(driver, url, retries=3):
    for attempt in range(1, retries + 1):
        try:
            if not is_driver_alive(driver):
                log.warning("  Driver dead, restarting...")
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(5)
                driver = create_driver()

            driver.get(url)
            time.sleep(random.uniform(5, 8))

            if "captcha" in driver.page_source[:3000].lower():
                log.warning(f"  CAPTCHA detected (attempt {attempt}/{retries}), waiting longer...")
                time.sleep(15)
                if "captcha" in driver.page_source[:3000].lower():
                    if attempt < retries:
                        log.warning("  Still captcha, restarting driver...")
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        time.sleep(5)
                        driver = create_driver()
                        continue
                    else:
                        return None, driver

            return driver.page_source, driver

        except Exception as e:
            log.error(f"  Page load error (attempt {attempt}): {e}")
            if attempt < retries:
                time.sleep(5)
            else:
                return None, driver

    return None, driver


# ── Load marketplace data ────────────────────────────────────────────────────
def load_marketplace_data():
    path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_marketplace.xlsx")
    if not os.path.exists(path):
        log.error(f"Marketplace file not found: {path}")
        return None, None

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None, None

    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = {}
        for i, h in enumerate(headers):
            record[h] = row[i] if i < len(row) else ""
        data.append(record)

    return headers, data


def get_unique_shops(data):
    shops = {}
    for row in data:
        shop = (row.get("Store Name") or "").strip()
        if not shop:
            continue
        state = (row.get("State") or "").strip()
        city = (row.get("City") or "").strip()
        if shop not in shops:
            shops[shop] = []
        if (state, city) not in shops[shop]:
            shops[shop].append((state, city))
    return shops


# ── Shop profile extraction ──────────────────────────────────────────────────
def extract_shop_profile(html, shop_name):
    soup = BeautifulSoup(html, "lxml")

    profile = {
        "shop_name": shop_name,
        "location": "",
        "rating": "",
        "icon_url": "",
        "shop_url": f"https://www.etsy.com/shop/{shop_name}",
    }

    loc_el = soup.find("p", class_="sb-shop-location")
    if loc_el:
        profile["location"] = loc_el.get_text(strip=True)

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, dict):
                agg = data.get("aggregateRating", {})
                if isinstance(agg, dict):
                    rv = agg.get("ratingValue")
                    if rv:
                        profile["rating"] = str(round(float(rv), 2))
                        break
        except (json.JSONDecodeError, AttributeError, ValueError):
            pass

    if not profile["rating"]:
        rating_input = soup.find("input", {"name": "rating"})
        if rating_input and rating_input.get("value"):
            try:
                profile["rating"] = str(round(float(rating_input["value"]), 2))
            except (ValueError, TypeError):
                pass

    icon_img = soup.find("img", class_="shop-icon-external")
    if icon_img:
        srcset = icon_img.get("srcset", "")
        if srcset:
            parts = [s.strip().split()[0] for s in srcset.split(",") if s.strip()]
            profile["icon_url"] = parts[-1] if parts else icon_img.get("src", "")
        else:
            profile["icon_url"] = icon_img.get("src", "")

    return profile


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Home and Garden — Etsy shop enricher")
    parser.add_argument("--test", action="store_true", help="Test mode: first 10 shops only")
    args = parser.parse_args()

    headers, data = load_marketplace_data()
    if not data:
        log.error("No marketplace data found. Run etsy_products.py first.")
        return

    log.info(f"Loaded {len(data)} marketplace rows")

    shops = get_unique_shops(data)
    shop_names = list(shops.keys())
    log.info(f"Found {len(shop_names)} unique shops")

    if args.test:
        shop_names = shop_names[:10]
        log.info("TEST MODE: processing first 10 shops")
        directory_rows, completed_shops = [], set()
    else:
        directory_rows, completed_shops = load_checkpoint()

    failed_shops = []
    shops_since_checkpoint = 0
    driver = create_driver()

    try:
        for idx, shop_name in enumerate(shop_names, 1):
            if shop_name in completed_shops:
                log.info(f"  [{idx}/{len(shop_names)}] Skipping {shop_name} (already done)")
                continue

            log.info(f"\n[{idx}/{len(shop_names)}] Visiting shop: {shop_name}")

            try:
                url = f"https://www.etsy.com/shop/{shop_name}"
                html, driver = fetch_page(driver, url)

                if not html:
                    log.warning(f"  Failed to load shop page for {shop_name}")
                    failed_shops.append(shop_name)
                    for state, city in shops[shop_name]:
                        directory_rows.append({
                            "state": state, "city": city,
                            "business_name": shop_name,
                            "address": f"{city}, {state}",
                            "phone": "N/A", "star_rating": "",
                            "website": "N/A", "profile_url": url,
                            "image_url": "", "source": "Etsy",
                        })
                else:
                    profile = extract_shop_profile(html, shop_name)
                    log.info(f"  Location: {profile['location'] or 'N/A'}")
                    log.info(f"  Rating:   {profile['rating'] or 'N/A'}")
                    log.info(f"  Icon:     {'yes' if profile['icon_url'] else 'no'}")

                    for state, city in shops[shop_name]:
                        directory_rows.append({
                            "state": state, "city": city,
                            "business_name": shop_name,
                            "address": f"{city}, {state}",
                            "phone": "N/A",
                            "star_rating": profile["rating"],
                            "website": "N/A",
                            "profile_url": profile["shop_url"],
                            "image_url": profile["icon_url"],
                            "source": "Etsy",
                        })

                completed_shops.add(shop_name)
                shops_since_checkpoint += 1
                time.sleep(random.uniform(2.0, 4.0))

            except Exception as e:
                log.error(f"  Shop {shop_name} failed: {e} — skipping")
                failed_shops.append(shop_name)

            if not args.test and shops_since_checkpoint >= CHECKPOINT_EVERY:
                save_checkpoint(directory_rows, completed_shops)
                shops_since_checkpoint = 0

    except KeyboardInterrupt:
        log.warning("\nInterrupted by user. Saving collected data...")
        if not args.test:
            save_checkpoint(directory_rows, completed_shops)
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    if directory_rows:
        dir_path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_directory.xlsx")
        os.makedirs(CATEGORIES_DIR, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Etsy Shops"

        dir_headers = ["State", "City", "Business Name", "Address", "Phone",
                       "Star Rating", "Website", "Profile URL", "Business Image URL", "Source"]
        ws.append(dir_headers)

        for row in directory_rows:
            ws.append([
                row["state"],
                row["city"],
                row["business_name"],
                row["address"],
                row["phone"],
                row["star_rating"],
                row["website"],
                row["profile_url"],
                row["image_url"],
                row["source"],
            ])

        wb.save(dir_path)
        log.info(f"\nSaved {len(directory_rows)} directory rows to {dir_path}")
        if not args.test and os.path.exists(CHECKPOINT_FILE):
            os.remove(CHECKPOINT_FILE)
            log.info("Checkpoint file removed (run complete)")

    log.info(f"\n{'='*60}")
    log.info(f"DONE")
    log.info(f"  Directory rows: {len(directory_rows)} (from {len(shop_names)} unique shops)")
    if failed_shops:
        log.info(f"  Failed shops ({len(failed_shops)}):")
        for s in failed_shops:
            log.info(f"    {s}")
    log.info(f"{'='*60}")


if __name__ == "__main__":
    main()
