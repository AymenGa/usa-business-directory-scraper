import openpyxl

CITIES_ORDER = {
    "Alabama": 0, "Alaska": 1, "Arizona": 2, "Arkansas": 3, "California": 4,
    "Colorado": 5, "Connecticut": 6, "Delaware": 7, "Florida": 8, "Georgia": 9,
    "Hawaii": 10, "Idaho": 11, "Illinois": 12, "Indiana": 13, "Iowa": 14,
    "Kansas": 15, "Kentucky": 16, "Louisiana": 17, "Maine": 18, "Maryland": 19,
    "Massachusetts": 20, "Michigan": 21, "Minnesota": 22, "Mississippi": 23,
    "Missouri": 24, "Montana": 25, "Nebraska": 26, "Nevada": 27,
    "New Hampshire": 28, "New Jersey": 29, "New Mexico": 30, "New York": 31,
    "North Carolina": 32, "North Dakota": 33, "Ohio": 34, "Oklahoma": 35,
    "Oregon": 36, "Pennsylvania": 37, "Rhode Island": 38, "South Carolina": 39,
    "South Dakota": 40, "Tennessee": 41, "Texas": 42, "Utah": 43,
    "Vermont": 44, "Virginia": 45, "Washington": 46, "West Virginia": 47,
    "Wisconsin": 48, "Wyoming": 49,
}

CITY_ORDER = {
    "Alabama": ["Huntsville","Birmingham","Montgomery","Mobile","Tuscaloosa"],
    "Alaska": ["Anchorage","Fairbanks","Juneau","Knik-Fairview","Badger"],
    "Arizona": ["Phoenix","Tucson","Mesa","Chandler","Gilbert"],
    "Arkansas": ["Little Rock","Fayetteville","Fort Smith","Springdale","Jonesboro"],
    "California": ["Los Angeles","San Diego","San Jose","San Francisco","Fresno"],
    "Colorado": ["Denver","Colorado Springs","Aurora","Fort Collins","Lakewood"],
    "Connecticut": ["Bridgeport","Stamford","New Haven","Hartford","Waterbury"],
    "Delaware": ["Wilmington","Dover","Newark","Middletown","Smyrna"],
    "Florida": ["Jacksonville","Miami","Tampa","Orlando","St. Petersburg"],
    "Georgia": ["Atlanta","Columbus","Augusta","Macon","Savannah"],
    "Hawaii": ["Honolulu","East Honolulu","Pearl City","Hilo","Kailua-Kona"],
    "Idaho": ["Boise","Meridian","Nampa","Idaho Falls","Caldwell"],
    "Illinois": ["Chicago","Aurora","Naperville","Joliet","Rockford"],
    "Indiana": ["Indianapolis","Fort Wayne","Evansville","South Bend","Carmel"],
    "Iowa": ["Des Moines","Cedar Rapids","Davenport","Sioux City","Iowa City"],
    "Kansas": ["Wichita","Overland Park","Kansas City","Olathe","Topeka"],
    "Kentucky": ["Louisville","Lexington","Bowling Green","Owensboro","Covington"],
    "Louisiana": ["New Orleans","Baton Rouge","Shreveport","Lafayette","Lake Charles"],
    "Maine": ["Portland","Lewiston","Bangor","South Portland","Auburn"],
    "Maryland": ["Baltimore","Frederick","Gaithersburg","Rockville","Bowie"],
    "Massachusetts": ["Boston","Worcester","Springfield","Cambridge","Lowell"],
    "Michigan": ["Detroit","Grand Rapids","Warren","Sterling Heights","Ann Arbor"],
    "Minnesota": ["Minneapolis","St. Paul","Rochester","Bloomington","Duluth"],
    "Mississippi": ["Jackson","Gulfport","Southaven","Hattiesburg","Biloxi"],
    "Missouri": ["Kansas City","St. Louis","Springfield","Columbia","Independence"],
    "Montana": ["Billings","Missoula","Great Falls","Bozeman","Butte"],
    "Nebraska": ["Omaha","Lincoln","Bellevue","Grand Island","Kearney"],
    "Nevada": ["Las Vegas","Henderson","North Las Vegas","Reno","Enterprise"],
    "New Hampshire": ["Manchester","Nashua","Concord","Derry","Dover"],
    "New Jersey": ["Newark","Jersey City","Paterson","Elizabeth","Lakewood"],
    "New Mexico": ["Albuquerque","Las Cruces","Rio Rancho","Santa Fe","Roswell"],
    "New York": ["New York City","Buffalo","Yonkers","Rochester","Syracuse"],
    "North Carolina": ["Charlotte","Raleigh","Greensboro","Durham","Winston-Salem"],
    "North Dakota": ["Fargo","Bismarck","Grand Forks","Minot","West Fargo"],
    "Ohio": ["Columbus","Cleveland","Cincinnati","Toledo","Akron"],
    "Oklahoma": ["Oklahoma City","Tulsa","Norman","Broken Arrow","Lawton"],
    "Oregon": ["Portland","Eugene","Salem","Gresham","Hillsboro"],
    "Pennsylvania": ["Philadelphia","Pittsburgh","Allentown","Reading","Erie"],
    "Rhode Island": ["Providence","Cranston","Warwick","Pawtucket","East Providence"],
    "South Carolina": ["Charleston","Columbia","North Charleston","Mount Pleasant","Rock Hill"],
    "South Dakota": ["Sioux Falls","Rapid City","Aberdeen","Brookings","Watertown"],
    "Tennessee": ["Nashville","Memphis","Knoxville","Chattanooga","Clarksville"],
    "Texas": ["Houston","San Antonio","Dallas","Fort Worth","Austin"],
    "Utah": ["Salt Lake City","West Valley City","West Jordan","Provo","St. George"],
    "Vermont": ["Burlington","South Burlington","Colchester","Rutland","Essex Junction"],
    "Virginia": ["Virginia Beach","Chesapeake","Arlington","Norfolk","Richmond"],
    "Washington": ["Seattle","Spokane","Tacoma","Vancouver","Bellevue"],
    "West Virginia": ["Charleston","Huntington","Morgantown","Parkersburg","Wheeling"],
    "Wisconsin": ["Milwaukee","Madison","Green Bay","Kenosha","Racine"],
    "Wyoming": ["Cheyenne","Casper","Gillette","Laramie","Rock Springs"],
}

wb = openpyxl.load_workbook("output/categories/electronics_gadgets_yelp_directory.xlsx")
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
data = rows[1:]

def sort_key(row):
    state = str(row[0] or "")
    city = str(row[1] or "")
    state_idx = CITIES_ORDER.get(state, 99)
    city_list = CITY_ORDER.get(state, [])
    city_idx = city_list.index(city) if city in city_list else 99
    return (state_idx, city_idx)

data.sort(key=sort_key)

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Yelp Businesses"
ws2.append(list(headers))
for row in data:
    ws2.append(list(row))
wb2.save("output/categories/electronics_gadgets_yelp_directory.xlsx")

print(f"Sorted {len(data)} rows")
print(f"First 3 rows: {[r[0]+'|'+r[1] for r in data[:3]]}")
print(f"Texas starts at row: {next((i+2 for i,r in enumerate(data) if r[0]=='Texas'), 'N/A')}")
