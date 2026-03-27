"""
Fashion / Accessories — Yelp business scraper (directory data).
Scrapes Yelp search results for fashion/accessories businesses.
3 businesses per city × 250 cities = 750 rows.
Visits individual business pages when phone/website missing from search.

Run:
  python scraper/fashion/yelp.py --test        # first 5 cities
  python scraper/fashion/yelp.py --state TX    # one state only
  python scraper/fashion/yelp.py               # all 250 cities
"""

import argparse
import json
import os
import re
import time
import random
import logging
import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import quote_plus

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CATEGORIES_DIR = os.path.join(BASE_DIR, "output", "categories")
METADATA_FILE  = os.path.join(CATEGORIES_DIR, "metadata.json")
PROGRESS_FILE  = os.path.join(CATEGORIES_DIR, "yelp_progress.json")

CATEGORY_SLUG  = "fashion_accessories"
BUSINESSES_PER_CITY = 3

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Category filters ─────────────────────────────────────────────────────────
# Categories that belong to Fashion / Accessories
ALLOWED_CATEGORIES = {
    "accessories", "women's clothing", "men's clothing", "fashion",
    "jewelry", "shoe stores", "shoes", "hats", "handbags", "sunglasses",
    "leather goods", "watches", "lingerie", "bridal", "formal wear",
    "vintage clothing", "clothing", "boutique", "eyewear",
    "women's accessories", "men's accessories",
}

# Categories to reject even if "accessories" is present
REJECT_CATEGORIES = {
    "costumes", "costume shop", "thrift stores", "gift shops",
    "sports wear", "sporting goods", "piercing", "tattoo",
    "auto parts", "auto accessories", "cell phone accessories",
    "phone repair", "computer repair", "pet stores",
    "home decor", "furniture", "hardware stores",
    "halloween", "dance", "dancewear", "dance supply",
    "party supplies", "uniform", "uniforms", "surplus",
}

# Business names that signal non-fashion even if categories look OK
REJECT_NAME_KEYWORDS = {
    "halloween", "spirit halloween", "costume", "dance depot",
    "dancewear", "party city", "uniform", "surplus",
    "goodwill", "salvation army", "dollar",
}

# ── 250 Cities (50 states × 5 cities) ────────────────────────────────────────
CITIES = {
    "Alabama": ["Huntsville", "Birmingham", "Montgomery", "Mobile", "Tuscaloosa"],
    "Alaska": ["Anchorage", "Fairbanks", "Juneau", "Knik-Fairview", "Badger"],
    "Arizona": ["Phoenix", "Tucson", "Mesa", "Chandler", "Gilbert"],
    "Arkansas": ["Little Rock", "Fayetteville", "Fort Smith", "Springdale", "Jonesboro"],
    "California": ["Los Angeles", "San Diego", "San Jose", "San Francisco", "Fresno"],
    "Colorado": ["Denver", "Colorado Springs", "Aurora", "Fort Collins", "Lakewood"],
    "Connecticut": ["Bridgeport", "Stamford", "New Haven", "Hartford", "Waterbury"],
    "Delaware": ["Wilmington", "Dover", "Newark", "Middletown", "Smyrna"],
    "Florida": ["Jacksonville", "Miami", "Tampa", "Orlando", "St. Petersburg"],
    "Georgia": ["Atlanta", "Columbus", "Augusta", "Macon", "Savannah"],
    "Hawaii": ["Honolulu", "East Honolulu", "Pearl City", "Hilo", "Kailua-Kona"],
    "Idaho": ["Boise", "Meridian", "Nampa", "Idaho Falls", "Caldwell"],
    "Illinois": ["Chicago", "Aurora", "Naperville", "Joliet", "Rockford"],
    "Indiana": ["Indianapolis", "Fort Wayne", "Evansville", "South Bend", "Carmel"],
    "Iowa": ["Des Moines", "Cedar Rapids", "Davenport", "Sioux City", "Iowa City"],
    "Kansas": ["Wichita", "Overland Park", "Kansas City", "Olathe", "Topeka"],
    "Kentucky": ["Louisville", "Lexington", "Bowling Green", "Owensboro", "Covington"],
    "Louisiana": ["New Orleans", "Baton Rouge", "Shreveport", "Lafayette", "Lake Charles"],
    "Maine": ["Portland", "Lewiston", "Bangor", "South Portland", "Auburn"],
    "Maryland": ["Baltimore", "Frederick", "Gaithersburg", "Rockville", "Bowie"],
    "Massachusetts": ["Boston", "Worcester", "Springfield", "Cambridge", "Lowell"],
    "Michigan": ["Detroit", "Grand Rapids", "Warren", "Sterling Heights", "Ann Arbor"],
    "Minnesota": ["Minneapolis", "St. Paul", "Rochester", "Bloomington", "Duluth"],
    "Mississippi": ["Jackson", "Gulfport", "Southaven", "Hattiesburg", "Biloxi"],
    "Missouri": ["Kansas City", "St. Louis", "Springfield", "Columbia", "Independence"],
    "Montana": ["Billings", "Missoula", "Great Falls", "Bozeman", "Butte"],
    "Nebraska": ["Omaha", "Lincoln", "Bellevue", "Grand Island", "Kearney"],
    "Nevada": ["Las Vegas", "Henderson", "North Las Vegas", "Reno", "Enterprise"],
    "New Hampshire": ["Manchester", "Nashua", "Concord", "Derry", "Dover"],
    "New Jersey": ["Newark", "Jersey City", "Paterson", "Elizabeth", "Lakewood"],
    "New Mexico": ["Albuquerque", "Las Cruces", "Rio Rancho", "Santa Fe", "Roswell"],
    "New York": ["New York City", "Buffalo", "Yonkers", "Rochester", "Syracuse"],
    "North Carolina": ["Charlotte", "Raleigh", "Greensboro", "Durham", "Winston-Salem"],
    "North Dakota": ["Fargo", "Bismarck", "Grand Forks", "Minot", "West Fargo"],
    "Ohio": ["Columbus", "Cleveland", "Cincinnati", "Toledo", "Akron"],
    "Oklahoma": ["Oklahoma City", "Tulsa", "Norman", "Broken Arrow", "Lawton"],
    "Oregon": ["Portland", "Eugene", "Salem", "Gresham", "Hillsboro"],
    "Pennsylvania": ["Philadelphia", "Pittsburgh", "Allentown", "Reading", "Erie"],
    "Rhode Island": ["Providence", "Cranston", "Warwick", "Pawtucket", "East Providence"],
    "South Carolina": ["Charleston", "Columbia", "North Charleston", "Mount Pleasant", "Rock Hill"],
    "South Dakota": ["Sioux Falls", "Rapid City", "Aberdeen", "Brookings", "Watertown"],
    "Tennessee": ["Nashville", "Memphis", "Knoxville", "Chattanooga", "Clarksville"],
    "Texas": ["Houston", "San Antonio", "Dallas", "Fort Worth", "Austin"],
    "Utah": ["Salt Lake City", "West Valley City", "West Jordan", "Provo", "St. George"],
    "Vermont": ["Burlington", "South Burlington", "Colchester", "Rutland", "Essex Junction"],
    "Virginia": ["Virginia Beach", "Chesapeake", "Arlington", "Norfolk", "Richmond"],
    "Washington": ["Seattle", "Spokane", "Tacoma", "Vancouver", "Bellevue"],
    "West Virginia": ["Charleston", "Huntington", "Morgantown", "Parkersburg", "Wheeling"],
    "Wisconsin": ["Milwaukee", "Madison", "Green Bay", "Kenosha", "Racine"],
    "Wyoming": ["Cheyenne", "Casper", "Gillette", "Laramie", "Rock Springs"],
}


# ── Browser helpers ──────────────────────────────────────────────────────────
def create_driver():
    """Create an undetected Chrome driver."""
    import undetected_chromedriver as uc
    options = uc.ChromeOptions()
    options.add_argument("--lang=en-US")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = uc.Chrome(options=options, headless=False, version_main=146)
    driver.set_page_load_timeout(30)
    return driver


def is_driver_alive(driver):
    """Check if the Chrome driver is still responsive."""
    try:
        _ = driver.title
        return True
    except Exception:
        return False


def fetch_page(driver, url, retries=3):
    """Navigate to URL and return page source. Retries on captcha/failure."""
    for attempt in range(1, retries + 1):
        # If driver is dead, recreate it before trying
        if not is_driver_alive(driver):
            log.warning(f"  Driver dead, recreating... (attempt {attempt})")
            try:
                driver.quit()
            except Exception:
                pass
            time.sleep(5)
            driver = create_driver()

        try:
            driver.get(url)
            wait = random.uniform(4, 7)
            time.sleep(wait)

            page_start = driver.page_source[:3000].lower()
            if "captcha" in page_start or "unusual traffic" in page_start:
                log.warning(f"  CAPTCHA/block detected (attempt {attempt}/{retries}), waiting...")
                time.sleep(20)
                page_start = driver.page_source[:3000].lower()
                if "captcha" in page_start or "unusual traffic" in page_start:
                    if attempt < retries:
                        log.warning("  Still blocked, restarting driver...")
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        time.sleep(10)
                        driver = create_driver()
                        continue
                    else:
                        return None, driver

            return driver.page_source, driver

        except Exception as e:
            log.error(f"  Page load error (attempt {attempt}): {e}")
            if attempt < retries:
                # Recreate driver on crash/timeout
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(5)
                driver = create_driver()
            else:
                # Last attempt failed — recreate driver so caller gets a live one
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(5)
                driver = create_driver()
                return None, driver

    return None, driver


# ── Category matching ────────────────────────────────────────────────────────
def is_fashion_business(categories_text):
    """Check if the business categories match Fashion / Accessories."""
    if not categories_text:
        return False
    cats = [c.strip().lower() for c in categories_text.split(",")]
    # Reject if any reject category is present
    # Use partial match only when reject keyword appears IN the business category
    # (not the other way — "accessories" should NOT match "auto accessories")
    for c in cats:
        if c in REJECT_CATEGORIES:
            return False
        for rej in REJECT_CATEGORIES:
            if rej in c:
                return False
    # Accept if any allowed category is present
    for c in cats:
        if c in ALLOWED_CATEGORIES:
            return True
    # Partial match — check if any allowed keyword appears in any category
    for c in cats:
        for allowed in ALLOWED_CATEGORIES:
            if allowed in c or c in allowed:
                return True
    return False


# ── Search page extraction ───────────────────────────────────────────────────
def extract_businesses_from_search(html, state, city):
    """Extract business cards from Yelp search results page."""
    soup = BeautifulSoup(html, "lxml")
    businesses = []

    # Yelp search results are in list items with specific patterns
    # Look for business cards — they contain name, rating, categories, address
    cards = soup.select('[data-testid="serp-ia-card"]')
    if not cards:
        # Fallback: find result containers
        cards = soup.select('div[class*="container"] li[class*="border"]')
    if not cards:
        # Another fallback: look for links to /biz/
        cards = soup.find_all("div", class_=re.compile(r"arrange-unit"))

    # More robust: find all business links and work from there
    biz_links = soup.find_all("a", href=re.compile(r"^/biz/[^?]+"))

    seen_urls = set()
    for link in biz_links:
        href = link.get("href", "")
        # Clean the URL
        biz_url = href.split("?")[0]
        full_url = f"https://www.yelp.com{biz_url}"

        if biz_url in seen_urls:
            continue

        # Get the business name from the link text
        name = link.get_text(strip=True)
        if not name or len(name) < 2 or len(name) > 100:
            continue
        # Skip numbered prefixes like "1. " or "2. "
        name = re.sub(r"^\d+\.\s*", "", name)
        if not name:
            continue
        # Skip non-business links like "more" or "Write a Review"
        if name.lower() in ("more", "write a review", "read more", "see all"):
            continue

        seen_urls.add(biz_url)

        # Walk up to the containing card to find more info
        card = link
        for _ in range(10):
            parent = card.parent
            if parent is None:
                break
            card = parent
            # Stop at a reasonable container size
            card_text = card.get_text(" ", strip=True)
            if len(card_text) > 200:
                break

        card_text = card.get_text(" ", strip=True) if card else ""

        # Extract rating from aria-label or text
        rating = ""
        rating_el = card.find(attrs={"aria-label": re.compile(r"\d[\d.]*\s*star")}) if card else None
        if rating_el:
            m = re.search(r"([\d.]+)\s*star", rating_el.get("aria-label", ""))
            if m:
                rating = m.group(1)

        # Extract categories
        categories = ""
        # Categories are often in spans near the rating
        cat_spans = card.find_all("span", class_=re.compile(r"css-")) if card else []
        for span in cat_spans:
            text = span.get_text(strip=True)
            # Categories look like "Accessories, Women's Clothing" or just "Jewelry"
            # Match multi-category strings with comma
            if "," in text and any(kw in text.lower() for kw in
                ["accessor", "cloth", "fashion", "jewel", "shoe", "hat",
                 "boutique", "wear", "leather", "watch", "eyewear", "bridal",
                 "lingerie", "sunglass", "handbag"]):
                categories = text
                break
            # Also match single-category spans (no comma needed)
            text_lower = text.lower().strip()
            if not categories and 2 < len(text_lower) < 40 and text_lower in ALLOWED_CATEGORIES:
                categories = text
                # Don't break — keep looking for a richer multi-category span

        # Extract address
        address = ""
        addr_candidates = card.find_all(string=re.compile(rf"{re.escape(city)}|{re.escape(state)}",
                                                           re.IGNORECASE)) if card else []
        for ac in addr_candidates:
            text = ac.strip()
            if len(text) > 5 and len(text) < 200:
                address = text
                break

        # Extract phone (usually not on search page but try)
        phone = ""
        phone_match = re.search(r"\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}", card_text)
        if phone_match:
            phone = phone_match.group(0)

        # Extract image URL
        image_url = ""
        img = card.find("img", src=re.compile(r"yelp")) if card else None
        if not img:
            img = card.find("img", src=re.compile(r"https://")) if card else None
        if img:
            image_url = img.get("src", "")

        businesses.append({
            "name": name,
            "rating": rating,
            "categories": categories,
            "address": address,
            "phone": phone,
            "website": "",
            "profile_url": full_url,
            "image_url": image_url,
        })

    return businesses


# ── Individual business page extraction ──────────────────────────────────────
def extract_business_details(html):
    """Extract phone, website, and categories from an individual Yelp business page."""
    soup = BeautifulSoup(html, "lxml")
    details = {"phone": "", "website": "", "address": "", "rating": "", "image_url": "", "categories": ""}

    # Phone — look for tel: links or phone pattern
    phone_link = soup.find("a", href=re.compile(r"^tel:"))
    if phone_link:
        details["phone"] = phone_link.get_text(strip=True)
    else:
        # Try finding phone in aside or business info sections
        phone_pattern = re.compile(r"\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}")
        for el in soup.find_all(["p", "span", "div"]):
            text = el.get_text(strip=True)
            m = phone_pattern.search(text)
            if m and len(text) < 50:
                details["phone"] = m.group(0)
                break

    # Website — look for "Business website" link or external URL links
    for link in soup.find_all("a", href=re.compile(r"biz_redir")):
        href = link.get("href", "")
        text = link.get_text(strip=True)
        if text and "http" not in text.lower() and len(text) < 80:
            details["website"] = text
            break
        # Extract URL from redirect
        url_match = re.search(r"url=([^&]+)", href)
        if url_match:
            from urllib.parse import unquote
            details["website"] = unquote(url_match.group(1))
            break

    if not details["website"]:
        # Look for external website links
        website_el = soup.find("a", string=re.compile(r"business.*website|visit.*website", re.I))
        if website_el:
            details["website"] = website_el.get("href", "")

    # Address
    addr_el = soup.find("address")
    if addr_el:
        details["address"] = addr_el.get_text(" ", strip=True)
    else:
        # Try JSON-LD
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string)
                if isinstance(data, dict) and "address" in data:
                    addr = data["address"]
                    if isinstance(addr, dict):
                        parts = [addr.get("streetAddress", ""),
                                 addr.get("addressLocality", ""),
                                 addr.get("addressRegion", "")]
                        details["address"] = ", ".join(p for p in parts if p)
                        break
            except (json.JSONDecodeError, AttributeError):
                pass

    # Rating from JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, dict):
                agg = data.get("aggregateRating", {})
                if isinstance(agg, dict) and agg.get("ratingValue"):
                    details["rating"] = str(round(float(agg["ratingValue"]), 2))
                    break
        except (json.JSONDecodeError, AttributeError, ValueError):
            pass

    # Image
    og_img = soup.find("meta", property="og:image")
    if og_img:
        details["image_url"] = og_img.get("content", "")

    # Categories — from JSON-LD or breadcrumb links on detail page
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, dict) and data.get("@type") == "LocalBusiness":
                # Yelp JSON-LD sometimes has category info
                if "additionalType" in data:
                    cats = data["additionalType"]
                    if isinstance(cats, list):
                        details["categories"] = ", ".join(cats)
                    elif isinstance(cats, str):
                        details["categories"] = cats
                    break
        except (json.JSONDecodeError, AttributeError):
            pass

    # Fallback: extract categories from page breadcrumbs/category links
    if not details["categories"]:
        cat_links = soup.find_all("a", href=re.compile(r"/search\?.*cflt="))
        if cat_links:
            cat_names = []
            for cl in cat_links:
                text = cl.get_text(strip=True)
                if text and len(text) < 50:
                    cat_names.append(text)
            if cat_names:
                details["categories"] = ", ".join(cat_names)

    # Another fallback: look for span tags that look like category labels
    if not details["categories"]:
        # On Yelp biz pages, categories appear as linked spans near the top
        cat_spans = soup.select('span[class*="css-"] a[href*="/search?"]')
        if cat_spans:
            cat_names = [s.get_text(strip=True) for s in cat_spans if s.get_text(strip=True)]
            if cat_names:
                details["categories"] = ", ".join(cat_names[:5])

    return details


# ── Progress tracking ────────────────────────────────────────────────────────
def load_progress():
    """Load set of completed city keys ('State|City')."""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_progress(done_set):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(sorted(done_set), f)


# ── Load existing Yelp directory rows ────────────────────────────────────────
def load_existing_yelp_rows():
    """Load previously scraped Yelp rows from the directory file."""
    path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_yelp_directory.xlsx")
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = {}
        for i, h in enumerate(headers):
            record[h] = row[i] if i < len(row) else ""
        data.append(record)
    return data


def save_yelp_rows(rows):
    """Save Yelp directory rows to intermediate file."""
    path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_yelp_directory.xlsx")
    os.makedirs(CATEGORIES_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yelp Businesses"

    headers = ["State", "City", "Business Name", "Address", "Phone",
               "Star Rating", "Website", "Profile URL", "Business Image URL", "Source"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("State", ""),
            row.get("City", ""),
            row.get("Business Name", ""),
            row.get("Address", ""),
            row.get("Phone", ""),
            row.get("Star Rating", ""),
            row.get("Website", ""),
            row.get("Profile URL", ""),
            row.get("Business Image URL", ""),
            row.get("Source", "Yelp"),
        ])

    wb.save(path)
    log.info(f"  Saved {len(rows)} Yelp rows to {path}")


# ── Merge Yelp + Etsy into final directory ────────────────────────────────────
def merge_directory():
    """Merge Yelp rows with existing Etsy rows into the final directory file."""
    etsy_path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_directory.xlsx")
    yelp_path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_yelp_directory.xlsx")
    final_path = etsy_path  # Overwrite the directory file

    headers = ["State", "City", "Business Name", "Address", "Phone",
               "Star Rating", "Website", "Profile URL", "Business Image URL", "Source"]

    all_rows = []

    # Load existing Etsy rows (Source = "Etsy")
    if os.path.exists(etsy_path):
        wb = openpyxl.load_workbook(etsy_path, read_only=True)
        ws = wb.active
        file_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(file_rows) > 1:
            file_headers = [str(h) for h in file_rows[0]]
            for row in file_rows[1:]:
                record = {}
                for i, h in enumerate(file_headers):
                    record[h] = row[i] if i < len(row) else ""
                # Keep only Etsy rows (don't duplicate Yelp rows if re-running)
                if record.get("Source", "") == "Etsy":
                    all_rows.append(record)

    # Load Yelp rows
    if os.path.exists(yelp_path):
        wb = openpyxl.load_workbook(yelp_path, read_only=True)
        ws = wb.active
        file_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(file_rows) > 1:
            file_headers = [str(h) for h in file_rows[0]]
            for row in file_rows[1:]:
                record = {}
                for i, h in enumerate(file_headers):
                    record[h] = row[i] if i < len(row) else ""
                all_rows.append(record)

    # Sort by State → City so Etsy + Yelp businesses for the same city are grouped
    all_rows.sort(key=lambda r: (str(r.get("State", "")).lower(), str(r.get("City", "")).lower()))

    # Write merged file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Directory"
    ws.append(headers)

    for record in all_rows:
        ws.append([record.get(h, "") for h in headers])

    wb.save(final_path)
    etsy_count = sum(1 for r in all_rows if r.get("Source") == "Etsy")
    yelp_count = sum(1 for r in all_rows if r.get("Source") == "Yelp")
    log.info(f"Merged directory: {etsy_count} Etsy + {yelp_count} Yelp = {len(all_rows)} total → {final_path}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Fashion / Accessories — Yelp scraper")
    parser.add_argument("--test", action="store_true", help="Test mode: first 5 cities only")
    parser.add_argument("--state", type=str, help="Scrape only one state (e.g., TX)")
    parser.add_argument("--merge-only", action="store_true", help="Only merge Etsy + Yelp, no scraping")
    args = parser.parse_args()

    if args.merge_only:
        merge_directory()
        return

    # Build city list
    city_list = []
    for state, cities in CITIES.items():
        if args.state:
            # Match state name or abbreviation
            if args.state.lower() not in (state.lower(), state[:2].lower()):
                continue
        for city in cities:
            city_list.append((state, city))

    if args.test:
        city_list = city_list[:5]
        log.info(f"TEST MODE: first 5 cities")

    log.info(f"Will scrape {len(city_list)} cities, {BUSINESSES_PER_CITY} businesses each")

    # Load progress and existing data
    done = load_progress()
    all_rows = load_existing_yelp_rows()
    log.info(f"Resuming: {len(done)} cities done, {len(all_rows)} rows collected")

    # Filter out already-done cities
    remaining = [(s, c) for s, c in city_list if f"{s}|{c}" not in done]
    log.info(f"Remaining: {len(remaining)} cities to scrape")

    if not remaining:
        log.info("All cities already scraped. Run --merge-only to rebuild directory.")
        merge_directory()
        return

    driver = create_driver()
    biz_pages_visited = 0

    try:
        for idx, (state, city) in enumerate(remaining, 1):
            key = f"{state}|{city}"
            log.info(f"\n[{idx}/{len(remaining)}] {city}, {state}")

            try:
                # Build search URL and paginate until we have enough businesses
                query = quote_plus("fashion accessories")
                location = quote_plus(f"{city}, {state}")
                base_search_url = f"https://www.yelp.com/search?find_desc={query}&find_loc={location}"

                MAX_PAGES = 4
                selected = []
                seen_profile_urls = set()

                for page in range(MAX_PAGES):
                    if len(selected) >= BUSINESSES_PER_CITY:
                        break

                    search_url = base_search_url if page == 0 else f"{base_search_url}&start={page * 10}"
                    if page > 0:
                        log.info(f"  Page {page + 1} (need {BUSINESSES_PER_CITY - len(selected)} more)")

                    html, driver = fetch_page(driver, search_url)
                    if not html:
                        log.warning(f"  Failed to load search page {page + 1} for {city}, {state}")
                        break

                    # Debug: save first city's first page HTML for inspection
                    if idx == 1 and page == 0:
                        debug_path = os.path.join(CATEGORIES_DIR, "_debug_yelp_search.html")
                        with open(debug_path, "w", encoding="utf-8") as f:
                            f.write(html)
                        log.info(f"  [DEBUG] Saved search HTML to {debug_path}")

                    # Extract businesses from search results
                    businesses = extract_businesses_from_search(html, state, city)
                    log.info(f"  Found {len(businesses)} raw results (page {page + 1})")

                    if not businesses:
                        log.info(f"  No more results, stopping pagination")
                        break

                    # Pre-filter by name (cheap — no page visits)
                    candidates = []
                    for biz in businesses:
                        if len(candidates) + len(selected) >= BUSINESSES_PER_CITY * 3:
                            break  # Enough candidates to try
                        if biz["profile_url"] in seen_profile_urls:
                            continue
                        seen_profile_urls.add(biz["profile_url"])

                        # Reject by business name first
                        name_lower = biz["name"].lower()
                        if any(rej in name_lower for rej in REJECT_NAME_KEYWORDS):
                            log.info(f"  Rejected (name): {biz['name']}")
                            continue

                        # Quick reject by search-page categories if clearly wrong
                        if biz["categories"]:
                            if not is_fashion_business(biz["categories"]):
                                log.info(f"  Rejected (search cats): {biz['name']} ({biz['categories']})")
                                continue

                        candidates.append(biz)

                    if not candidates:
                        log.info(f"  No candidates on page {page + 1}")
                        if page < MAX_PAGES - 1:
                            time.sleep(random.uniform(2, 4))
                        continue

                    # Visit each candidate's detail page to verify categories + get phone/website
                    for biz in candidates:
                        if len(selected) >= BUSINESSES_PER_CITY:
                            break

                        log.info(f"  → Visiting: {biz['name']}")
                        detail_html, driver = fetch_page(driver, biz["profile_url"])
                        biz_pages_visited += 1

                        if detail_html:
                            details = extract_business_details(detail_html)

                            # Update business info
                            if details["phone"]:
                                biz["phone"] = details["phone"]
                            if details["website"]:
                                biz["website"] = details["website"]
                            if details["address"]:
                                biz["address"] = details["address"]
                            if details["rating"]:
                                biz["rating"] = details["rating"]
                            if details["image_url"]:
                                biz["image_url"] = details["image_url"]

                            # VALIDATE with detail-page categories (the real ones)
                            real_cats = details.get("categories", "")
                            if real_cats:
                                biz["categories"] = real_cats
                                if not is_fashion_business(real_cats):
                                    log.info(f"    REJECTED (detail cats): {real_cats}")
                                    time.sleep(random.uniform(2, 4))
                                    continue
                                log.info(f"    VERIFIED: {real_cats}")
                            else:
                                # No categories on detail page either — accept if name looks right
                                name_lower = biz["name"].lower()
                                fashion_name_hints = ["fashion", "accessor", "jewel", "boutique", "cloth",
                                                      "shoe", "hat", "leather", "eyewear", "bridal", "sunglass"]
                                if not any(hint in name_lower for hint in fashion_name_hints):
                                    log.info(f"    SKIPPED (no categories anywhere): {biz['name']}")
                                    time.sleep(random.uniform(2, 4))
                                    continue
                                log.info(f"    Included (name match, no cats): {biz['name']}")

                            # Require both phone AND website — skip incomplete businesses
                            if not biz["phone"] or not biz["website"]:
                                log.info(f"    SKIPPED (incomplete): Phone={biz['phone'] or 'MISSING'} Web={biz['website'] or 'MISSING'}")
                                time.sleep(random.uniform(2, 4))
                                continue

                            # Reject Yelp placeholder images — only keep real business photos
                            if biz.get("image_url") and "yelp_og_image" in biz["image_url"]:
                                biz["image_url"] = ""

                            log.info(f"    Phone: {biz['phone']} | Web: {biz['website']}")

                        selected.append(biz)
                        time.sleep(random.uniform(2, 4))

                    # Small delay between pagination
                    if page < MAX_PAGES - 1 and len(selected) < BUSINESSES_PER_CITY:
                        time.sleep(random.uniform(2, 4))

                log.info(f"  Selected {len(selected)} verified businesses")

                # Build directory rows
                for biz in selected:
                    # Clean address — remove "Closed until ..." prefix
                    address = biz.get("address", "")
                    address = re.sub(r"(Open|Closed)\s+(until|now).*?(AM|PM)\s*", "", address, flags=re.I).strip()
                    if not address:
                        address = f"{city}, {state}"

                    all_rows.append({
                        "State": state,
                        "City": city,
                        "Business Name": biz["name"],
                        "Address": address,
                        "Phone": biz["phone"] or "N/A",
                        "Star Rating": biz["rating"],
                        "Website": biz["website"] or "N/A",
                        "Profile URL": biz["profile_url"],
                        "Business Image URL": biz["image_url"],
                        "Source": "Yelp",
                    })

            except Exception as e:
                log.error(f"  CITY ERROR ({city}, {state}): {e}")
                log.info(f"  Recovering — recreating driver and continuing...")
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(10)
                driver = create_driver()

            # Mark done and save periodically (runs for both success and error)
            done.add(key)
            if idx % 5 == 0 or idx == len(remaining):
                save_progress(done)
                save_yelp_rows(all_rows)
                log.info(f"  [Checkpoint] {len(done)} cities done, {len(all_rows)} total rows")

            # Delay between cities
            time.sleep(random.uniform(3, 6))

    except KeyboardInterrupt:
        log.warning("\nInterrupted by user. Saving progress...")
    finally:
        save_progress(done)
        save_yelp_rows(all_rows)
        try:
            driver.quit()
        except Exception:
            pass

    # ── Summary ──────────────────────────────────────────────────────────────
    log.info(f"\n{'='*60}")
    log.info(f"DONE")
    log.info(f"  Cities scraped: {len(done)}")
    log.info(f"  Yelp rows: {len(all_rows)}")
    log.info(f"  Business pages visited: {biz_pages_visited}")
    log.info(f"{'='*60}")

    # Merge with Etsy data
    merge_directory()


if __name__ == "__main__":
    main()
