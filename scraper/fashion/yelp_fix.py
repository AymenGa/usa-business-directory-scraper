"""
Fashion / Accessories — Yelp wrong-city replacement scraper.

Reads fashion_accessories_directory.xlsx, finds Yelp rows where the address
does not match the City column, then scrapes Yelp to find correct-city
replacements. Swaps bad rows in-place and saves the updated Excel.

Run:
  python scraper/fashion/yelp_fix.py --test    # first 5 affected cities only
  python scraper/fashion/yelp_fix.py           # all 113 affected cities
"""

import argparse
import json
import os
import re
import time
import random
import logging
import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import quote_plus

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CATEGORIES_DIR = os.path.join(BASE_DIR, "output", "categories")
DIRECTORY_FILE = os.path.join(CATEGORIES_DIR, "fashion_accessories_directory.xlsx")
PROGRESS_FILE  = os.path.join(CATEGORIES_DIR, "yelp_fix_progress.json")

MAX_PAGES = 5

CITY_ALIASES = {
    "new york city": "new york",
    "east honolulu": "honolulu",
    "knik-fairview": "wasilla",
    "enterprise": "las vegas",
    "badger": "badger",  # use zip code override below
}

# Override the Yelp search location string for cities Yelp doesn't recognise
LOCATION_OVERRIDES = {
    "badger": "99705",  # Badger, AK zip — Yelp ignores "Badger, Alaska"
}

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Category filters (copied from yelp.py) ───────────────────────────────────
ALLOWED_CATEGORIES = {
    "accessories", "women's clothing", "men's clothing", "fashion",
    "jewelry", "shoe stores", "shoes", "hats", "handbags", "sunglasses",
    "leather goods", "watches", "lingerie", "bridal", "formal wear",
    "vintage clothing", "clothing", "boutique", "eyewear",
    "women's accessories", "men's accessories",
}

REJECT_CATEGORIES = {
    "costumes", "costume shop", "thrift stores", "gift shops",
    "sports wear", "sporting goods", "piercing", "tattoo",
    "auto parts", "auto accessories", "cell phone accessories",
    "phone repair", "computer repair", "pet stores",
    "home decor", "furniture", "hardware stores",
    "halloween", "dance", "dancewear", "dance supply",
    "party supplies", "uniform", "uniforms", "surplus",
}

REJECT_NAME_KEYWORDS = {
    "halloween", "spirit halloween", "costume", "dance depot",
    "dancewear", "party city", "uniform", "surplus",
    "goodwill", "salvation army", "dollar",
}


# ── Category check ────────────────────────────────────────────────────────────
def is_fashion_business(categories_text):
    if not categories_text:
        return False
    cats = [c.strip().lower() for c in categories_text.split(",")]
    for c in cats:
        if c in REJECT_CATEGORIES:
            return False
        for rej in REJECT_CATEGORIES:
            if rej in c:
                return False
    for c in cats:
        if c in ALLOWED_CATEGORIES:
            return True
    for c in cats:
        for allowed in ALLOWED_CATEGORIES:
            if allowed in c or c in allowed:
                return True
    return False


# ── Browser helpers (copied from yelp.py) ────────────────────────────────────
def create_driver():
    import undetected_chromedriver as uc
    options = uc.ChromeOptions()
    options.add_argument("--lang=en-US")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = uc.Chrome(options=options, headless=False, version_main=146)
    driver.set_page_load_timeout(30)
    return driver


def is_driver_alive(driver):
    try:
        _ = driver.title
        return True
    except Exception:
        return False


def fetch_page(driver, url, retries=3):
    for attempt in range(1, retries + 1):
        if not is_driver_alive(driver):
            log.warning(f"  Driver dead, recreating... (attempt {attempt})")
            try:
                driver.quit()
            except Exception:
                pass
            time.sleep(5)
            driver = create_driver()
        try:
            driver.get(url)
            time.sleep(random.uniform(4, 7))
            page_start = driver.page_source[:3000].lower()
            if "captcha" in page_start or "unusual traffic" in page_start:
                log.warning(f"  CAPTCHA detected (attempt {attempt}/{retries}), waiting...")
                time.sleep(20)
                page_start = driver.page_source[:3000].lower()
                if "captcha" in page_start or "unusual traffic" in page_start:
                    if attempt < retries:
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        time.sleep(10)
                        driver = create_driver()
                        continue
                    else:
                        return None, driver
            return driver.page_source, driver
        except Exception as e:
            log.error(f"  Page load error (attempt {attempt}): {e}")
            try:
                driver.quit()
            except Exception:
                pass
            time.sleep(5)
            driver = create_driver()
            if attempt == retries:
                return None, driver
    return None, driver


# ── Search page extraction (copied from yelp.py) ─────────────────────────────
def extract_businesses_from_search(html, state, city):
    soup = BeautifulSoup(html, "lxml")
    biz_links = soup.find_all("a", href=re.compile(r"^/biz/[^?]+"))
    seen_urls = set()
    businesses = []

    for link in biz_links:
        href = link.get("href", "")
        biz_url = href.split("?")[0]
        full_url = f"https://www.yelp.com{biz_url}"
        if biz_url in seen_urls:
            continue
        name = link.get_text(strip=True)
        if not name or len(name) < 2 or len(name) > 100:
            continue
        name = re.sub(r"^\d+\.\s*", "", name)
        if not name:
            continue
        if name.lower() in ("more", "write a review", "read more", "see all"):
            continue
        seen_urls.add(biz_url)

        card = link
        for _ in range(10):
            parent = card.parent
            if parent is None:
                break
            card = parent
            if len(card.get_text(" ", strip=True)) > 200:
                break

        card_text = card.get_text(" ", strip=True) if card else ""

        rating = ""
        rating_el = card.find(attrs={"aria-label": re.compile(r"\d[\d.]*\s*star")}) if card else None
        if rating_el:
            m = re.search(r"([\d.]+)\s*star", rating_el.get("aria-label", ""))
            if m:
                rating = m.group(1)

        categories = ""
        cat_spans = card.find_all("span", class_=re.compile(r"css-")) if card else []
        for span in cat_spans:
            text = span.get_text(strip=True)
            if "," in text and any(kw in text.lower() for kw in
                ["accessor", "cloth", "fashion", "jewel", "shoe", "hat",
                 "boutique", "wear", "leather", "watch", "eyewear", "bridal",
                 "lingerie", "sunglass", "handbag"]):
                categories = text
                break
            text_lower = text.lower().strip()
            if not categories and 2 < len(text_lower) < 40 and text_lower in ALLOWED_CATEGORIES:
                categories = text

        address = ""
        addr_candidates = card.find_all(
            string=re.compile(rf"{re.escape(city)}|{re.escape(state)}", re.IGNORECASE)
        ) if card else []
        for ac in addr_candidates:
            text = ac.strip()
            if 5 < len(text) < 200:
                address = text
                break

        image_url = ""
        img = card.find("img", src=re.compile(r"yelp")) if card else None
        if not img:
            img = card.find("img", src=re.compile(r"https://")) if card else None
        if img:
            image_url = img.get("src", "")

        businesses.append({
            "name": name,
            "rating": rating,
            "categories": categories,
            "address": address,
            "phone": "",
            "website": "",
            "profile_url": full_url,
            "image_url": image_url,
        })

    return businesses


# ── Detail page extraction (copied from yelp.py) ─────────────────────────────
def extract_business_details(html):
    soup = BeautifulSoup(html, "lxml")
    details = {"phone": "", "website": "", "address": "", "rating": "", "image_url": "", "categories": ""}

    phone_link = soup.find("a", href=re.compile(r"^tel:"))
    if phone_link:
        details["phone"] = phone_link.get_text(strip=True)
    else:
        phone_pattern = re.compile(r"\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}")
        for el in soup.find_all(["p", "span", "div"]):
            text = el.get_text(strip=True)
            m = phone_pattern.search(text)
            if m and len(text) < 50:
                details["phone"] = m.group(0)
                break

    for link in soup.find_all("a", href=re.compile(r"biz_redir")):
        href = link.get("href", "")
        text = link.get_text(strip=True)
        if text and "http" not in text.lower() and len(text) < 80:
            details["website"] = text
            break
        url_match = re.search(r"url=([^&]+)", href)
        if url_match:
            from urllib.parse import unquote
            details["website"] = unquote(url_match.group(1))
            break

    if not details["website"]:
        website_el = soup.find("a", string=re.compile(r"business.*website|visit.*website", re.I))
        if website_el:
            details["website"] = website_el.get("href", "")

    addr_el = soup.find("address")
    if addr_el:
        details["address"] = addr_el.get_text(" ", strip=True)
    else:
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data = json.loads(script.string)
                if isinstance(data, dict) and "address" in data:
                    addr = data["address"]
                    if isinstance(addr, dict):
                        parts = [addr.get("streetAddress", ""),
                                 addr.get("addressLocality", ""),
                                 addr.get("addressRegion", "")]
                        details["address"] = ", ".join(p for p in parts if p)
                        break
            except (json.JSONDecodeError, AttributeError):
                pass

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, dict):
                agg = data.get("aggregateRating", {})
                if isinstance(agg, dict) and agg.get("ratingValue"):
                    details["rating"] = str(round(float(agg["ratingValue"]), 2))
                    break
        except (json.JSONDecodeError, AttributeError, ValueError):
            pass

    og_img = soup.find("meta", property="og:image")
    if og_img:
        details["image_url"] = og_img.get("content", "")

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, dict) and data.get("@type") == "LocalBusiness":
                if "additionalType" in data:
                    cats = data["additionalType"]
                    if isinstance(cats, list):
                        details["categories"] = ", ".join(cats)
                    elif isinstance(cats, str):
                        details["categories"] = cats
                    break
        except (json.JSONDecodeError, AttributeError):
            pass

    if not details["categories"]:
        cat_links = soup.find_all("a", href=re.compile(r"/search\?.*cflt="))
        if cat_links:
            cat_names = [cl.get_text(strip=True) for cl in cat_links if cl.get_text(strip=True)]
            if cat_names:
                details["categories"] = ", ".join(cat_names)

    if not details["categories"]:
        cat_spans = soup.select('span[class*="css-"] a[href*="/search?"]')
        if cat_spans:
            cat_names = [s.get_text(strip=True) for s in cat_spans if s.get_text(strip=True)]
            if cat_names:
                details["categories"] = ", ".join(cat_names[:5])

    return details


# ── Progress ──────────────────────────────────────────────────────────────────
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def save_progress(done_set):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(list(done_set), f)


# ── Excel helpers ─────────────────────────────────────────────────────────────
def load_excel():
    """Load the directory Excel and return (workbook, worksheet, rows as list of dicts)."""
    wb = openpyxl.load_workbook(DIRECTORY_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return wb, ws, headers, rows


def find_bad_rows(rows):
    """Return dict: (state, city) -> list of row indices (0-based in rows list) that are wrong-city."""
    from collections import defaultdict
    bad = defaultdict(list)
    for i, row in enumerate(rows):
        source = str(row.get("Source", "")).strip().lower()
        if "yelp" not in source:
            continue
        state = str(row.get("State", "")).strip()
        city = str(row.get("City", "")).strip()
        address = str(row.get("Address", "")).strip()
        if not address:
            continue
        city_lower = city.lower()
        check_city = CITY_ALIASES.get(city_lower, city_lower)
        if check_city not in address.lower():
            bad[(state, city)].append(i)
    return bad


def good_names_for_city(rows, state, city):
    """Return set of business names already correctly placed in this city."""
    names = set()
    for row in rows:
        if row.get("State") == state and row.get("City") == city:
            source = str(row.get("Source", "")).strip().lower()
            if "yelp" in source:
                addr = str(row.get("Address", "")).lower()
                city_lower = city.lower()
                check_city = CITY_ALIASES.get(city_lower, city_lower)
                if check_city in addr:
                    names.add(str(row.get("Business Name", "")).strip().lower())
    return names


def global_seen_profile_urls(rows):
    """Return set of all Yelp profile URLs already in the Excel (across all cities)."""
    urls = set()
    for row in rows:
        source = str(row.get("Source", "")).strip().lower()
        if "yelp" in source:
            url = str(row.get("Profile URL", "")).strip()
            if url:
                urls.add(url.split("?")[0])  # strip query params
    return urls


def write_excel(ws, headers, rows):
    """Overwrite worksheet rows (keeping header) with updated rows list."""
    # Clear existing data rows
    for row_idx in range(ws.max_row, 1, -1):
        ws.delete_rows(row_idx)
    # Write updated rows
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--test", action="store_true", help="Process first 5 affected cities only")
    args = parser.parse_args()

    log.info("Loading Excel...")
    wb, ws, headers, rows = load_excel()

    bad_by_city = find_bad_rows(rows)
    city_list = sorted(bad_by_city.keys())  # (state, city)
    log.info(f"Found {sum(len(v) for v in bad_by_city.values())} bad rows across {len(city_list)} cities")

    done = load_progress()
    remaining = [(s, c) for s, c in city_list if f"{s}|{c}" not in done]

    if args.test:
        remaining = remaining[:5]
        log.info(f"TEST MODE — processing first 5 cities: {[c for _, c in remaining]}")

    if not remaining:
        log.info("Nothing to fix — all cities already processed.")
        return

    # Build global URL set — no business already in the Excel can be used as a replacement
    global_urls = global_seen_profile_urls(rows)
    log.info(f"Global dedup pool: {len(global_urls)} existing Yelp URLs")

    driver = create_driver()

    try:
        for idx, (state, city) in enumerate(remaining, 1):
            key = f"{state}|{city}"
            bad_indices = bad_by_city[(state, city)]
            needed = len(bad_indices)
            log.info(f"\n[{idx}/{len(remaining)}] {city}, {state} — need {needed} replacement(s)")

            # Names already correctly in this city (to avoid duplicates)
            existing_names = good_names_for_city(rows, state, city)
            # Also exclude names of the bad rows themselves (we're replacing them)
            bad_names = {str(rows[i].get("Business Name", "")).strip().lower() for i in bad_indices}
            skip_names = existing_names - bad_names

            replacements = []
            query = quote_plus("fashion accessories boutique clothing jewelry")
            loc_str = LOCATION_OVERRIDES.get(city.lower(), f"{city}, {state}")
            location = quote_plus(loc_str)

            try:
                for page in range(MAX_PAGES):
                    if len(replacements) >= needed:
                        break

                    start = page * 10
                    url = (f"https://www.yelp.com/search?find_desc={query}"
                           f"&find_loc={location}&start={start}")
                    log.info(f"  Page {page + 1}: {url}")

                    html, driver = fetch_page(driver, url)
                    if not html:
                        log.warning(f"  Failed to load page {page + 1}")
                        continue

                    candidates = extract_businesses_from_search(html, state, city)
                    log.info(f"  Found {len(candidates)} candidates on page {page + 1}")

                    for biz in candidates:
                        if len(replacements) >= needed:
                            break

                        # Skip if name already in this city
                        if biz["name"].lower() in skip_names:
                            log.info(f"  Skipped (duplicate): {biz['name']}")
                            continue

                        # Quick category reject
                        if biz["categories"] and not is_fashion_business(biz["categories"]):
                            log.info(f"  Skipped (bad cats): {biz['name']}")
                            continue

                        # Name reject
                        if any(rej in biz["name"].lower() for rej in REJECT_NAME_KEYWORDS):
                            log.info(f"  Skipped (name): {biz['name']}")
                            continue

                        log.info(f"  -> Visiting: {biz['name']}")
                        detail_html, driver = fetch_page(driver, biz["profile_url"])

                        if not detail_html:
                            log.warning(f"    Detail page failed, skipping")
                            time.sleep(random.uniform(2, 4))
                            continue

                        details = extract_business_details(detail_html)

                        if details["phone"]:
                            biz["phone"] = details["phone"]
                        if details["website"]:
                            biz["website"] = details["website"]
                        if details["address"]:
                            biz["address"] = details["address"]
                        if details["rating"]:
                            biz["rating"] = details["rating"]
                        if details["image_url"]:
                            biz["image_url"] = details["image_url"]

                        # Verify address is in the correct city
                        addr_lower = biz["address"].lower()
                        city_lower = city.lower()
                        check_city = CITY_ALIASES.get(city_lower, city_lower)
                        if addr_lower and check_city not in addr_lower:
                            log.info(f"    REJECTED (wrong city): {biz['address']}")
                            time.sleep(random.uniform(2, 4))
                            continue

                        # Validate categories from detail page
                        real_cats = details.get("categories", "")
                        if real_cats:
                            biz["categories"] = real_cats
                            if not is_fashion_business(real_cats):
                                log.info(f"    REJECTED (detail cats): {real_cats}")
                                time.sleep(random.uniform(2, 4))
                                continue

                        # Require phone (website optional)
                        if not biz["phone"]:
                            log.info(f"    SKIPPED (no phone): {biz['name']}")
                            time.sleep(random.uniform(2, 4))
                            continue

                        # Global dedup — reject if this URL is already used anywhere in the Excel
                        clean_url = biz["profile_url"].split("?")[0]
                        if clean_url in global_urls:
                            log.info(f"    SKIPPED (already in Excel elsewhere): {biz['name']}")
                            time.sleep(random.uniform(2, 4))
                            continue

                        # Clean placeholder images
                        if biz.get("image_url") and "yelp_og_image" in biz["image_url"]:
                            biz["image_url"] = ""

                        log.info(f"    ACCEPTED: {biz['name']} | {biz['address']}")
                        replacements.append(biz)
                        skip_names.add(biz["name"].lower())
                        global_urls.add(clean_url)  # reserve so later cities can't reuse it
                        time.sleep(random.uniform(2, 4))

                    if page < MAX_PAGES - 1 and len(replacements) < needed:
                        time.sleep(random.uniform(2, 4))

                log.info(f"  Found {len(replacements)} replacement(s) for {needed} needed")

                # Swap bad rows with replacements
                for slot, bad_idx in enumerate(bad_indices):
                    if slot < len(replacements):
                        rep = replacements[slot]
                        address = rep.get("address", "")
                        address = re.sub(r"(Open|Closed)\s+(until|now).*?(AM|PM)\s*", "", address, flags=re.I).strip()
                        if not address:
                            address = f"{city}, {state}"
                        rows[bad_idx] = {
                            "State": state,
                            "City": city,
                            "Business Name": rep["name"],
                            "Address": address,
                            "Phone": rep["phone"] or "N/A",
                            "Star Rating": rep["rating"],
                            "Website": rep["website"] or "N/A",
                            "Profile URL": rep["profile_url"],
                            "Business Image URL": rep["image_url"],
                            "Source": "Yelp",
                        }
                        log.info(f"  Replaced: '{rows[bad_idx]['Business Name']}' -> '{rep['name']}'")
                    else:
                        # Not enough replacements found — keep original but log it
                        log.warning(f"  No replacement found for slot {slot + 1} in {city}, {state} — keeping original")

            except Exception as e:
                log.error(f"  CITY ERROR ({city}, {state}): {e}")
                try:
                    driver.quit()
                except Exception:
                    pass
                time.sleep(10)
                driver = create_driver()

            done.add(key)
            save_progress(done)

            # Save Excel after every city
            log.info(f"  Saving Excel...")
            write_excel(ws, headers, rows)
            wb.save(DIRECTORY_FILE)
            log.info(f"  Saved.")

    except KeyboardInterrupt:
        log.info("\nInterrupted — saving progress and Excel...")
        write_excel(ws, headers, rows)
        wb.save(DIRECTORY_FILE)
        save_progress(done)

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    # Final save
    write_excel(ws, headers, rows)
    wb.save(DIRECTORY_FILE)
    save_progress(done)
    log.info(f"\nDone. Excel saved to {DIRECTORY_FILE}")


if __name__ == "__main__":
    main()
