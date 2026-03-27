"""
Fashion / Accessories — Etsy shop profile enricher (directory data).
Visits each unique shop page to extract location, rating, icon.
Creates Directory rows (10 columns) from Etsy shop profiles.

Run:
  python scraper/fashion/etsy_shops.py --test   # first 10 shops
  python scraper/fashion/etsy_shops.py          # all shops
"""

import argparse
import json
import os
import time
import random
import logging
import openpyxl
from bs4 import BeautifulSoup

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CATEGORIES_DIR = os.path.join(BASE_DIR, "output", "categories")
METADATA_FILE  = os.path.join(CATEGORIES_DIR, "metadata.json")

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Fashion / Accessories config ─────────────────────────────────────────────
CATEGORY_NAME = "Fashion / Accessories"
CATEGORY_SLUG = "fashion_accessories"
SOURCE        = "www.etsy.com"


# ── Browser helpers ──────────────────────────────────────────────────────────
def create_driver():
    """Create an undetected Chrome driver."""
    import undetected_chromedriver as uc
    options = uc.ChromeOptions()
    options.add_argument("--lang=en-US")
    options.add_argument("--disable-blink-features=AutomationControlled")
    driver = uc.Chrome(options=options, headless=False, version_main=146)
    driver.set_page_load_timeout(30)
    return driver


def fetch_page(driver, url, retries=3):
    """Navigate to URL and return page source. Retries on captcha/failure."""
    for attempt in range(1, retries + 1):
        try:
            driver.get(url)
            wait = random.uniform(5, 8)
            time.sleep(wait)

            if "captcha" in driver.page_source[:3000].lower():
                log.warning(f"  CAPTCHA detected (attempt {attempt}/{retries}), waiting longer...")
                time.sleep(15)
                if "captcha" in driver.page_source[:3000].lower():
                    if attempt < retries:
                        log.warning("  Still captcha, restarting driver...")
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        time.sleep(5)
                        driver = create_driver()
                        continue
                    else:
                        return None, driver

            return driver.page_source, driver

        except Exception as e:
            log.error(f"  Page load error (attempt {attempt}): {e}")
            if attempt < retries:
                time.sleep(5)
            else:
                return None, driver

    return None, driver


# ── Load marketplace data ────────────────────────────────────────────────────
def load_marketplace_data():
    """Load the existing marketplace xlsx. Returns list of dicts with original data."""
    path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_marketplace.xlsx")
    if not os.path.exists(path):
        log.error(f"Marketplace file not found: {path}")
        return None, None

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None, None

    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = {}
        for i, h in enumerate(headers):
            record[h] = row[i] if i < len(row) else ""
        data.append(record)

    return headers, data


def get_unique_shops(data):
    """Extract unique (shop_name, state, city) combos from marketplace data.
    Returns dict: shop_name -> list of (state, city) pairs where they appear."""
    shops = {}
    for row in data:
        shop = (row.get("Store Name") or "").strip()
        if not shop:
            continue
        state = (row.get("State") or "").strip()
        city = (row.get("City") or "").strip()
        if shop not in shops:
            shops[shop] = []
        if (state, city) not in shops[shop]:
            shops[shop].append((state, city))
    return shops


# ── Shop profile extraction ─────────────────────────────────────────────────
def extract_shop_profile(html, shop_name):
    """Extract shop info from an Etsy shop page."""
    soup = BeautifulSoup(html, "lxml")

    profile = {
        "shop_name": shop_name,
        "location": "",
        "rating": "",
        "icon_url": "",
        "shop_url": f"https://www.etsy.com/shop/{shop_name}",
    }

    # Location — <p class="sb-shop-location ...">Birmingham, Alabama</p>
    loc_el = soup.find("p", class_="sb-shop-location")
    if loc_el:
        profile["location"] = loc_el.get_text(strip=True)

    # Rating — shop star rating
    # Look for rating in structured data or page elements
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, dict):
                agg = data.get("aggregateRating", {})
                if isinstance(agg, dict):
                    rv = agg.get("ratingValue")
                    if rv:
                        profile["rating"] = str(round(float(rv), 2))
                        break
        except (json.JSONDecodeError, AttributeError, ValueError):
            pass

    if not profile["rating"]:
        # Try finding rating from input or span elements
        rating_input = soup.find("input", {"name": "rating"})
        if rating_input and rating_input.get("value"):
            try:
                profile["rating"] = str(round(float(rating_input["value"]), 2))
            except (ValueError, TypeError):
                pass

    # Shop icon — <img class="shop-icon-external ...">
    icon_img = soup.find("img", class_="shop-icon-external")
    if icon_img:
        # Prefer the largest srcset image, fall back to src
        srcset = icon_img.get("srcset", "")
        if srcset:
            # Pick the last (largest) entry: "url 100w, url 200w" -> url
            parts = [s.strip().split()[0] for s in srcset.split(",") if s.strip()]
            profile["icon_url"] = parts[-1] if parts else icon_img.get("src", "")
        else:
            profile["icon_url"] = icon_img.get("src", "")

    return profile


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Fashion / Accessories — Etsy shop enricher")
    parser.add_argument("--test", action="store_true", help="Test mode: first 10 shops only")
    args = parser.parse_args()

    # Load existing marketplace data
    headers, data = load_marketplace_data()
    if not data:
        log.error("No marketplace data found. Run etsy_products.py first.")
        return

    log.info(f"Loaded {len(data)} marketplace rows")

    # Get unique shops
    shops = get_unique_shops(data)
    shop_names = list(shops.keys())
    log.info(f"Found {len(shop_names)} unique shops")

    if args.test:
        shop_names = shop_names[:10]
        log.info(f"TEST MODE: processing first 10 shops")

    # Create browser and visit each shop
    driver = create_driver()
    directory_rows = []
    failed_shops = []

    try:
        for idx, shop_name in enumerate(shop_names, 1):
            log.info(f"\n[{idx}/{len(shop_names)}] Visiting shop: {shop_name}")

            url = f"https://www.etsy.com/shop/{shop_name}"
            html, driver = fetch_page(driver, url)

            if not html:
                log.warning(f"  Failed to load shop page for {shop_name}")
                failed_shops.append(shop_name)
                # Still create a row with minimal info
                for state, city in shops[shop_name]:
                    directory_rows.append({
                        "state": state,
                        "city": city,
                        "business_name": shop_name,
                        "address": f"{city}, {state}",
                        "phone": "N/A",
                        "star_rating": "",
                        "website": "N/A",
                        "profile_url": url,
                        "image_url": "",
                        "source": "Etsy",
                    })
                continue

            profile = extract_shop_profile(html, shop_name)
            log.info(f"  Location: {profile['location'] or 'N/A'}")
            log.info(f"  Rating: {profile['rating'] or 'N/A'}")
            log.info(f"  Icon: {'yes' if profile['icon_url'] else 'no'}")

            # Always use marketplace city/state for address (Etsy only shows state-level location)
            for state, city in shops[shop_name]:
                address = f"{city}, {state}"
                directory_rows.append({
                    "state": state,
                    "city": city,
                    "business_name": shop_name,
                    "address": address,
                    "phone": "N/A",
                    "star_rating": profile["rating"],
                    "website": "N/A",
                    "profile_url": profile["shop_url"],
                    "image_url": profile["icon_url"],
                    "source": "Etsy",
                })

            # Delay between shop visits
            time.sleep(random.uniform(2.0, 4.0))

    except KeyboardInterrupt:
        log.warning("\nInterrupted by user. Saving collected data...")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    # ── Save directory rows ──────────────────────────────────────────────────
    if directory_rows:
        dir_path = os.path.join(CATEGORIES_DIR, f"{CATEGORY_SLUG}_directory.xlsx")
        os.makedirs(CATEGORIES_DIR, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Etsy Shops"

        dir_headers = ["State", "City", "Business Name", "Address", "Phone",
                        "Star Rating", "Website", "Profile URL", "Business Image URL", "Source"]
        ws.append(dir_headers)

        for row in directory_rows:
            ws.append([
                row["state"],
                row["city"],
                row["business_name"],
                row["address"],
                row["phone"],
                row["star_rating"],
                row["website"],
                row["profile_url"],
                row["image_url"],
                row["source"],
            ])

        wb.save(dir_path)
        log.info(f"\nSaved {len(directory_rows)} directory rows to {dir_path}")

    # ── Summary ──────────────────────────────────────────────────────────────
    log.info(f"\n{'='*60}")
    log.info(f"DONE")
    log.info(f"  Directory rows: {len(directory_rows)} (from {len(shop_names)} unique shops)")
    if failed_shops:
        log.info(f"  Failed shops ({len(failed_shops)}):")
        for s in failed_shops:
            log.info(f"    {s}")
    log.info(f"{'='*60}")


if __name__ == "__main__":
    main()
